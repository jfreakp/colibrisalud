from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Paciente, Cita, NumeroNotificacion, Mensaje, Notificacion
from datetime import datetime
from django.utils import timezone
import io


@login_required(login_url='login')
def dashboard(request):
    return render(request, 'home/dashboard.html')


@login_required(login_url='login')
def pacientes_import(request):
    """Vista para importar pacientes desde Excel"""
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            messages.error(request, 'Por favor selecciona un archivo')
            return redirect('pacientes_import')
        
        archivo = request.FILES['archivo']
        
        try:
            # Leer el archivo Excel
            contenido = archivo.read()
            workbook = load_workbook(io.BytesIO(contenido))
            worksheet = workbook.active
            
            importados = 0
            actualizados = 0
            errores = 0
            errores_detalles = []
            
            # Iterar desde la fila 2 (fila 1 es encabezado)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nombre = row[0]
                    apellido = row[1]
                    movil = row[2]
                    
                    if not all([nombre, apellido, movil]):
                        errores += 1
                        errores_detalles.append(f"Fila {row_idx}: Faltan datos obligatorios")
                        continue
                    
                    # Limpiar y validar el móvil
                    movil = str(movil).strip()
                    
                    # Crear o actualizar paciente
                    paciente, creado = Paciente.objects.update_or_create(
                        movil=movil,
                        defaults={
                            'nombre': str(nombre).strip(),
                            'apellido': str(apellido).strip(),
                            'activo': True
                        }
                    )
                    
                    if creado:
                        importados += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores += 1
                    errores_detalles.append(f"Fila {row_idx}: {str(e)}")
            
            # Mensaje de éxito
            mensaje = f"✓ Importación completada: {importados} nuevos, {actualizados} actualizados, {errores} errores"
            messages.success(request, mensaje)
            
            if errores_detalles:
                for detalle in errores_detalles[:5]:  # Mostrar primeros 5 errores
                    messages.warning(request, detalle)
                if len(errores_detalles) > 5:
                    messages.warning(request, f"... y {len(errores_detalles) - 5} errores más")
            
            return redirect('pacientes_lista')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('pacientes_import')
    
    return render(request, 'home/pacientes_import.html')


@login_required(login_url='login')
def pacientes_lista(request):
    """Vista para listar pacientes"""
    pacientes = Paciente.objects.all()
    
    # Filtros
    filtro_estado = request.GET.get('estado', '')
    if filtro_estado:
        pacientes = pacientes.filter(activo=filtro_estado == 'activo')
    
    busqueda = request.GET.get('q', '')
    if busqueda:
        pacientes = pacientes.filter(
            nombre__icontains=busqueda) | pacientes.filter(
            apellido__icontains=busqueda) | pacientes.filter(
            movil__icontains=busqueda
        )
    
    contexto = {
        'pacientes': pacientes,
        'total': Paciente.objects.count(),
        'activos': Paciente.objects.filter(activo=True).count(),
        'inactivos': Paciente.objects.filter(activo=False).count(),
    }
    
    return render(request, 'home/pacientes_lista.html', contexto)


@login_required(login_url='login')
def paciente_toggle(request, pk):
    """Toggle para activar/desactivar paciente"""
    try:
        paciente = Paciente.objects.get(pk=pk)
        paciente.activo = not paciente.activo
        paciente.save()
        
        estado = "activado" if paciente.activo else "desactivado"
        messages.success(request, f'Paciente {estado} correctamente')
    except Paciente.DoesNotExist:
        messages.error(request, 'Paciente no encontrado')
    
    return redirect('pacientes_lista')

def descargar_plantilla_excel(request):
    """Descargar plantilla Excel para importar pacientes"""
    # Crear workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Pacientes"
    
    # Configurar encabezados
    headers = ['Nombre', 'Apellido', 'Móvil']
    header_fill = PatternFill(start_color="0C6E68", end_color="0C6E68", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar ejemplos
    ejemplos = [
        ['Juan', 'Pérez', '+593991234567'],
        ['María', 'González', '+593992345678'],
        ['Carlos', 'López', '+593993456789'],
    ]
    
    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col_num, valor in enumerate(ejemplo, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 25
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 20
    
    # Guardar en memoria
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_pacientes.xlsx"'
    
    workbook.save(response)
    return response


@login_required(login_url='login')
def citas_import(request):
    """Vista para importar citas desde Excel"""
    if request.method == 'POST':
        if 'archivo' not in request.FILES:
            messages.error(request, 'Por favor selecciona un archivo')
            return redirect('citas_import')
        
        archivo = request.FILES['archivo']
        
        try:
            # Leer el archivo Excel
            contenido = archivo.read()
            workbook = load_workbook(io.BytesIO(contenido))
            worksheet = workbook.active
            
            importadas = 0
            actualizadas = 0
            errores = 0
            errores_detalles = []
            
            # Iterar desde la fila 2 (fila 1 es encabezado)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    movil_paciente = row[0]
                    fecha_str = row[1]
                    hora_str = row[2]
                    estado = row[3] if len(row) > 3 and row[3] else 'pendiente'
                    
                    if not all([movil_paciente, fecha_str, hora_str]):
                        errores += 1
                        errores_detalles.append(f"Fila {row_idx}: Faltan datos obligatorios")
                        continue
                    
                    # Buscar paciente por móvil
                    movil_paciente = str(movil_paciente).strip()
                    try:
                        paciente = Paciente.objects.get(movil=movil_paciente)
                    except Paciente.DoesNotExist:
                        errores += 1
                        errores_detalles.append(f"Fila {row_idx}: Paciente con móvil {movil_paciente} no encontrado")
                        continue
                    
                    # Convertir fecha
                    if isinstance(fecha_str, datetime):
                        fecha = fecha_str.date()
                    else:
                        try:
                            fecha = datetime.strptime(str(fecha_str), '%Y-%m-%d').date()
                        except ValueError:
                            try:
                                fecha = datetime.strptime(str(fecha_str), '%d/%m/%Y').date()
                            except ValueError:
                                errores += 1
                                errores_detalles.append(f"Fila {row_idx}: Formato de fecha inválido")
                                continue
                    
                    # Convertir hora
                    if isinstance(hora_str, datetime):
                        hora = hora_str.time()
                    else:
                        try:
                            hora = datetime.strptime(str(hora_str), '%H:%M').time()
                        except ValueError:
                            try:
                                hora = datetime.strptime(str(hora_str), '%H:%M:%S').time()
                            except ValueError:
                                errores += 1
                                errores_detalles.append(f"Fila {row_idx}: Formato de hora inválido")
                                continue
                    
                    # Validar estado
                    estados_validos = ['pendiente', 'confirmada', 'cancelada', 'atendida']
                    estado = str(estado).lower().strip()
                    if estado not in estados_validos:
                        estado = 'pendiente'
                    
                    # Crear o actualizar cita
                    cita, creada = Cita.objects.update_or_create(
                        paciente=paciente,
                        fecha=fecha,
                        hora=hora,
                        defaults={
                            'estado': estado
                        }
                    )
                    
                    if creada:
                        importadas += 1
                    else:
                        actualizadas += 1
                        
                except Exception as e:
                    errores += 1
                    errores_detalles.append(f"Fila {row_idx}: {str(e)}")
            
            # Mensaje de éxito
            mensaje = f"✓ Importación completada: {importadas} nuevas, {actualizadas} actualizadas, {errores} errores"
            messages.success(request, mensaje)
            
            if errores_detalles:
                for detalle in errores_detalles[:5]:  # Mostrar primeros 5 errores
                    messages.warning(request, detalle)
                if len(errores_detalles) > 5:
                    messages.warning(request, f"... y {len(errores_detalles) - 5} errores más")
            
            return redirect('citas_lista')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('citas_import')
    
    return render(request, 'home/citas_import.html')


@login_required(login_url='login')
def citas_lista(request):
    """Vista para listar citas"""
    citas = Cita.objects.select_related('paciente').all()
    
    # Filtros
    filtro_estado = request.GET.get('estado', '')
    if filtro_estado:
        citas = citas.filter(estado=filtro_estado)
    
    busqueda = request.GET.get('q', '')
    if busqueda:
        citas = citas.filter(
            paciente__nombre__icontains=busqueda) | citas.filter(
            paciente__apellido__icontains=busqueda) | citas.filter(
            paciente__movil__icontains=busqueda
        )
    
    contexto = {
        'citas': citas,
        'total': Cita.objects.count(),
        'pendientes': Cita.objects.filter(estado='pendiente').count(),
        'confirmadas': Cita.objects.filter(estado='confirmada').count(),
        'atendidas': Cita.objects.filter(estado='atendida').count(),
        'canceladas': Cita.objects.filter(estado='cancelada').count(),
    }
    
    return render(request, 'home/citas_lista.html', contexto)


@login_required(login_url='login')
def descargar_plantilla_citas(request):
    """Descargar plantilla Excel para importar citas"""
    # Crear workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Citas"
    
    # Configurar encabezados
    headers = ['Móvil Paciente', 'Fecha', 'Hora', 'Estado']
    header_fill = PatternFill(start_color="0C6E68", end_color="0C6E68", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar ejemplos
    ejemplos = [
        ['+593991234567', '2026-01-20', '09:00', 'pendiente'],
        ['+593992345678', '2026-01-20', '10:30', 'confirmada'],
        ['+593993456789', '2026-01-21', '14:00', 'pendiente'],
    ]
    
    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col_num, valor in enumerate(ejemplo, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = valor
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar ancho de columnas
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 15
    
    # Guardar en memoria
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_citas.xlsx"'
    
    workbook.save(response)
    return response


@login_required(login_url='login')
def numeros_lista(request):
    """Vista para listar números de notificación"""
    numeros = NumeroNotificacion.objects.all()
    
    # Filtros
    filtro_estado = request.GET.get('estado', '')
    if filtro_estado:
        numeros = numeros.filter(activo=filtro_estado == 'activo')
    
    busqueda = request.GET.get('q', '')
    if busqueda:
        numeros = numeros.filter(
            numero__icontains=busqueda) | numeros.filter(
            descripcion__icontains=busqueda
        )
    
    contexto = {
        'numeros': numeros,
        'total': NumeroNotificacion.objects.count(),
        'activos': NumeroNotificacion.objects.filter(activo=True).count(),
        'inactivos': NumeroNotificacion.objects.filter(activo=False).count(),
    }
    
    return render(request, 'home/numeros_lista.html', contexto)


@login_required(login_url='login')
def numero_crear(request):
    """Vista para crear un número de notificación"""
    if request.method == 'POST':
        numero = request.POST.get('numero', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        
        if not numero:
            messages.error(request, 'El número es obligatorio')
            return redirect('numero_crear')
        
        if NumeroNotificacion.objects.filter(numero=numero).exists():
            messages.error(request, 'Este número ya existe')
            return redirect('numero_crear')
        
        try:
            NumeroNotificacion.objects.create(
                numero=numero,
                descripcion=descripcion,
                activo=True
            )
            messages.success(request, f'Número {numero} creado correctamente')
            return redirect('numeros_lista')
        except Exception as e:
            messages.error(request, f'Error al crear el número: {str(e)}')
            return redirect('numero_crear')
    
    return render(request, 'home/numero_form.html', {'titulo': 'Crear Número'})


@login_required(login_url='login')
def numero_editar(request, pk):
    """Vista para editar un número de notificación"""
    numero = get_object_or_404(NumeroNotificacion, pk=pk)
    
    if request.method == 'POST':
        numero.descripcion = request.POST.get('descripcion', '').strip()
        numero.activo = request.POST.get('activo') == 'on'
        
        try:
            numero.save()
            messages.success(request, 'Número actualizado correctamente')
            return redirect('numeros_lista')
        except Exception as e:
            messages.error(request, f'Error al actualizar el número: {str(e)}')
            return redirect('numero_editar', pk=pk)
    
    contexto = {
        'numero': numero,
        'titulo': 'Editar Número'
    }
    return render(request, 'home/numero_form.html', contexto)


@login_required(login_url='login')
def numero_eliminar(request, pk):
    """Vista para eliminar un número de notificación"""
    numero = get_object_or_404(NumeroNotificacion, pk=pk)
    
    if request.method == 'POST':
        numero_str = numero.numero
        numero.delete()
        messages.success(request, f'Número {numero_str} eliminado correctamente')
        return redirect('numeros_lista')
    
    contexto = {'numero': numero}
    return render(request, 'home/numero_eliminar.html', contexto)


@login_required(login_url='login')
def numero_toggle(request, pk):
    """Toggle para activar/desactivar número"""
    try:
        numero = NumeroNotificacion.objects.get(pk=pk)
        numero.activo = not numero.activo
        numero.save()
        
        estado = "activado" if numero.activo else "desactivado"
        messages.success(request, f'Número {estado} correctamente')
    except NumeroNotificacion.DoesNotExist:
        messages.error(request, 'Número no encontrado')
    
    return redirect('numeros_lista')


@login_required(login_url='login')
def mensajes_lista(request):
    """Vista para listar mensajes"""
    mensajes_obj = Mensaje.objects.all()
    
    # Filtros
    filtro_tipo = request.GET.get('tipo', '')
    if filtro_tipo:
        mensajes_obj = mensajes_obj.filter(tipo=filtro_tipo)
    
    filtro_estado = request.GET.get('estado', '')
    if filtro_estado:
        mensajes_obj = mensajes_obj.filter(activo=filtro_estado == 'activo')
    
    busqueda = request.GET.get('q', '')
    if busqueda:
        mensajes_obj = mensajes_obj.filter(
            titulo__icontains=busqueda) | mensajes_obj.filter(
            contenido__icontains=busqueda
        )
    
    contexto = {
        'mensajes': mensajes_obj,
        'total': Mensaje.objects.count(),
        'activos': Mensaje.objects.filter(activo=True).count(),
        'inactivos': Mensaje.objects.filter(activo=False).count(),
        'tipos': Mensaje.TIPO_CHOICES,
    }
    
    return render(request, 'home/mensajes_lista.html', contexto)


@login_required(login_url='login')
def mensaje_crear(request):
    """Vista para crear un mensaje"""
    if request.method == 'POST':
        titulo = request.POST.get('titulo', '').strip()
        contenido = request.POST.get('contenido', '').strip()
        tipo = request.POST.get('tipo', 'otro').strip()
        
        if not titulo or not contenido:
            messages.error(request, 'Título y contenido son obligatorios')
            return redirect('mensaje_crear')
        
        try:
            Mensaje.objects.create(
                titulo=titulo,
                contenido=contenido,
                tipo=tipo,
                activo=True
            )
            messages.success(request, f'Mensaje "{titulo}" creado correctamente')
            return redirect('mensajes_lista')
        except Exception as e:
            messages.error(request, f'Error al crear el mensaje: {str(e)}')
            return redirect('mensaje_crear')
    
    contexto = {
        'titulo': 'Crear Mensaje',
        'tipos': Mensaje.TIPO_CHOICES,
    }
    return render(request, 'home/mensaje_form.html', contexto)


@login_required(login_url='login')
def mensaje_editar(request, pk):
    """Vista para editar un mensaje"""
    mensaje = get_object_or_404(Mensaje, pk=pk)
    
    if request.method == 'POST':
        mensaje.titulo = request.POST.get('titulo', '').strip()
        mensaje.contenido = request.POST.get('contenido', '').strip()
        mensaje.tipo = request.POST.get('tipo', 'otro').strip()
        mensaje.activo = request.POST.get('activo') == 'on'
        
        if not mensaje.titulo or not mensaje.contenido:
            messages.error(request, 'Título y contenido son obligatorios')
            return redirect('mensaje_editar', pk=pk)
        
        try:
            mensaje.save()
            messages.success(request, 'Mensaje actualizado correctamente')
            return redirect('mensajes_lista')
        except Exception as e:
            messages.error(request, f'Error al actualizar el mensaje: {str(e)}')
            return redirect('mensaje_editar', pk=pk)
    
    contexto = {
        'mensaje': mensaje,
        'titulo': 'Editar Mensaje',
        'tipos': Mensaje.TIPO_CHOICES,
    }
    return render(request, 'home/mensaje_form.html', contexto)


@login_required(login_url='login')
def mensaje_eliminar(request, pk):
    """Vista para eliminar un mensaje"""
    mensaje = get_object_or_404(Mensaje, pk=pk)
    
    if request.method == 'POST':
        titulo = mensaje.titulo
        mensaje.delete()
        messages.success(request, f'Mensaje "{titulo}" eliminado correctamente')
        return redirect('mensajes_lista')
    
    contexto = {'mensaje': mensaje}
    return render(request, 'home/mensaje_eliminar.html', contexto)


@login_required(login_url='login')
def mensaje_toggle(request, pk):
    """Toggle para activar/desactivar mensaje"""
    try:
        mensaje = Mensaje.objects.get(pk=pk)
        mensaje.activo = not mensaje.activo
        mensaje.save()
        
        estado = "activado" if mensaje.activo else "desactivado"
        messages.success(request, f'Mensaje {estado} correctamente')
    except Mensaje.DoesNotExist:
        messages.error(request, 'Mensaje no encontrado')
    
    return redirect('mensajes_lista')


@login_required(login_url='login')
def notificaciones_lista(request):
    """Vista para listar notificaciones"""
    notificaciones = Notificacion.objects.all()
    
    # Filtros
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    buscar = request.GET.get('buscar', '')
    
    if estado:
        notificaciones = notificaciones.filter(estado=estado)
    
    if fecha_desde:
        try:
            fecha = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            notificaciones = notificaciones.filter(fecha_programada__date__gte=fecha)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            notificaciones = notificaciones.filter(fecha_programada__date__lte=fecha)
        except ValueError:
            pass
    
    if buscar:
        notificaciones = notificaciones.filter(
            numero_destinatario__icontains=buscar
        ) | notificaciones.filter(
            cita__paciente__nombre__icontains=buscar
        ) | notificaciones.filter(
            cita__paciente__movil__icontains=buscar
        )
    
    # Estadísticas
    total = Notificacion.objects.count()
    enviadas = Notificacion.objects.filter(estado='enviada').count()
    pendientes = Notificacion.objects.filter(estado='pendiente').count()
    fallidas = Notificacion.objects.filter(estado='fallida').count()
    
    contexto = {
        'notificaciones': notificaciones,
        'estados': Notificacion.ESTADO_CHOICES,
        'total': total,
        'enviadas': enviadas,
        'pendientes': pendientes,
        'fallidas': fallidas,
    }
    return render(request, 'home/notificaciones_lista.html', contexto)


@login_required(login_url='login')
def notificacion_crear(request):
    """Vista para crear una notificación"""
    if request.method == 'POST':
        cita_id = request.POST.get('cita')
        numero_id = request.POST.get('numero')
        mensaje_id = request.POST.get('mensaje')
        fecha_programada_str = request.POST.get('fecha_programada')
        
        try:
            cita = Cita.objects.get(pk=cita_id)
            numero = NumeroNotificacion.objects.get(pk=numero_id, activo=True)
            mensaje = Mensaje.objects.get(pk=mensaje_id, activo=True)
            
            # Convertir fecha y hora
            fecha_programada = datetime.strptime(fecha_programada_str, '%Y-%m-%dT%H:%M')
            fecha_programada = timezone.make_aware(fecha_programada)
            
            # Validar que no sea fecha pasada
            if fecha_programada.date() < timezone.now().date():
                messages.error(request, 'No se puede programar notificaciones para fechas pasadas')
                return redirect('notificacion_crear')
            
            # Crear notificación
            notificacion = Notificacion.objects.create(
                cita=cita,
                numero=numero,
                mensaje=mensaje,
                numero_destinatario=cita.paciente.movil,
                numero_origen=numero.numero,
                fecha_programada=fecha_programada,
            )
            
            messages.success(request, 'Notificación creada correctamente')
            return redirect('notificaciones_lista')
        
        except (Cita.DoesNotExist, NumeroNotificacion.DoesNotExist, Mensaje.DoesNotExist) as e:
            messages.error(request, 'Error: verifique que los datos seleccionados existan y estén activos')
            return redirect('notificacion_crear')
        except ValueError:
            messages.error(request, 'Formato de fecha inválido')
            return redirect('notificacion_crear')
        except Exception as e:
            messages.error(request, f'Error al crear notificación: {str(e)}')
            return redirect('notificacion_crear')
    
    # GET - Mostrar formulario
    citas = Cita.objects.filter(estado__in=['pendiente', 'confirmada']).order_by('fecha', 'hora')
    numeros = NumeroNotificacion.objects.filter(activo=True)
    mensajes = Mensaje.objects.filter(activo=True)
    
    contexto = {
        'titulo': 'Crear Notificación',
        'citas': citas,
        'numeros': numeros,
        'mensajes': mensajes,
    }
    return render(request, 'home/notificacion_form.html', contexto)


@login_required(login_url='login')
def notificacion_editar(request, pk):
    """Vista para editar una notificación"""
    notificacion = get_object_or_404(Notificacion, pk=pk)
    
    if request.method == 'POST':
        try:
            # Solo permitir editar si es pendiente
            if notificacion.estado != 'pendiente':
                messages.error(request, 'Solo se pueden editar notificaciones pendientes')
                return redirect('notificacion_editar', pk=pk)
            
            numero_id = request.POST.get('numero')
            mensaje_id = request.POST.get('mensaje')
            fecha_programada_str = request.POST.get('fecha_programada')
            
            numero = NumeroNotificacion.objects.get(pk=numero_id, activo=True)
            mensaje = Mensaje.objects.get(pk=mensaje_id, activo=True)
            
            # Convertir fecha
            fecha_programada = datetime.strptime(fecha_programada_str, '%Y-%m-%dT%H:%M')
            fecha_programada = timezone.make_aware(fecha_programada)
            
            # Validar que no sea fecha pasada
            if fecha_programada.date() < timezone.now().date():
                messages.error(request, 'No se puede programar notificaciones para fechas pasadas')
                return redirect('notificacion_editar', pk=pk)
            
            # Actualizar
            notificacion.numero = numero
            notificacion.mensaje = mensaje
            notificacion.numero_origen = numero.numero
            notificacion.fecha_programada = fecha_programada
            notificacion.save()
            
            messages.success(request, 'Notificación actualizada correctamente')
            return redirect('notificaciones_lista')
        
        except Exception as e:
            messages.error(request, f'Error al actualizar notificación: {str(e)}')
            return redirect('notificacion_editar', pk=pk)
    
    # GET - Mostrar formulario
    numeros = NumeroNotificacion.objects.filter(activo=True)
    mensajes = Mensaje.objects.filter(activo=True)
    
    contexto = {
        'notificacion': notificacion,
        'titulo': 'Editar Notificación',
        'numeros': numeros,
        'mensajes': mensajes,
    }
    return render(request, 'home/notificacion_form.html', contexto)


@login_required(login_url='login')
def notificacion_eliminar(request, pk):
    """Vista para eliminar una notificación"""
    notificacion = get_object_or_404(Notificacion, pk=pk)
    
    if request.method == 'POST':
        if notificacion.estado != 'pendiente':
            messages.error(request, 'Solo se pueden eliminar notificaciones pendientes')
            return redirect('notificaciones_lista')
        
        numero_dest = notificacion.numero_destinatario
        notificacion.delete()
        messages.success(request, f'Notificación para {numero_dest} eliminada correctamente')
        return redirect('notificaciones_lista')
    
    contexto = {'notificacion': notificacion}
    return render(request, 'home/notificacion_eliminar.html', contexto)


@login_required(login_url='login')
def notificacion_enviar(request, pk):
    """Vista para enviar una notificación via WhatsApp"""
    notificacion = get_object_or_404(Notificacion, pk=pk)
    
    if notificacion.estado != 'pendiente':
        messages.error(request, f'Esta notificación ya ha sido {notificacion.get_estado_display().lower()}')
        return redirect('notificaciones_lista')
    
    # Validación de fecha
    if notificacion.fecha_programada.date() < timezone.now().date():
        messages.error(request, 'No se puede enviar notificaciones de fechas pasadas')
        notificacion.estado = 'fallida'
        notificacion.mensaje_error = 'Intento de envío de notificación con fecha pasada'
        notificacion.save()
        return redirect('notificaciones_lista')
    
    try:
        # TODO: Implementar integración con WhatsApp API
        # Por ahora, marcar como enviada (simulación)
        notificacion.estado = 'enviada'
        notificacion.fecha_enviada = timezone.now()
        notificacion.response_whatsapp = {
            'status': 'success',
            'message': 'Notificación enviada (simulada)',
            'timestamp': timezone.now().isoformat()
        }
        notificacion.save()
        
        messages.success(request, f'Notificación enviada a {notificacion.numero_destinatario}')
    except Exception as e:
        notificacion.estado = 'fallida'
        notificacion.mensaje_error = str(e)
        notificacion.response_whatsapp = {
            'status': 'error',
            'error': str(e),
            'timestamp': timezone.now().isoformat()
        }
        notificacion.save()
        
        messages.error(request, f'Error al enviar notificación: {str(e)}')
    
    return redirect('notificaciones_lista')
